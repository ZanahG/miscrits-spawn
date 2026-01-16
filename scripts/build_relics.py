import json
import re
from pathlib import Path

from openpyxl import load_workbook


def norm(s: str) -> str:
  s = (s or "").strip().lower()
  s = re.sub(r"\s+", " ", s)
  return s


def pick_header_map(headers):
  """
  Detecta columnas típicas sin depender de nombres exactos.
  Busca:
   - name / miscrit
   - relic1..relic4 (o relic 1..4)
  """
  h = [norm(x) for x in headers]

  def find_any(keys):
    for k in keys:
      if k in h:
        return h.index(k)
    return None

  name_idx = find_any(["name", "miscrit", "miscrit name", "miscrits", "miscrit_name"])

  # Relic columns: intenta varias variantes
  relic_idxs = []
  for i in range(1, 5):
    idx = find_any([f"relic{i}", f"relic {i}", f"best relic {i}", f"bestrelic{i}", f"slot{i}"])
    relic_idxs.append(idx)

  return name_idx, relic_idxs


def main():
  excel_path = Path(__file__).resolve().parent / "RelicsForMiscritByName.xlsx"
  out_path = Path(__file__).resolve().parent.parent / "assets" / "data" / "best_relics.json"

  if not excel_path.exists():
    raise FileNotFoundError(
      f"No encuentro el Excel en: {excel_path}\n"
      "Pon el archivo 'RelicsForMiscritByName.xlsx' dentro de la carpeta /scripts."
    )

  wb = load_workbook(excel_path, data_only=True)
  ws = wb[wb.sheetnames[0]]

  rows = list(ws.iter_rows(values_only=True))
  if not rows:
    raise RuntimeError("La hoja está vacía.")

  headers = [str(x or "").strip() for x in rows[0]]
  name_idx, relic_idxs = pick_header_map(headers)

  if name_idx is None:
    raise RuntimeError(
      "No pude detectar la columna del nombre del Miscrit.\n"
      "Asegúrate que exista una columna llamada: Name o Miscrit (o similar)."
    )

  if all(i is None for i in relic_idxs):
    raise RuntimeError(
      "No pude detectar columnas de reliquias.\n"
      "Asegúrate que existan columnas tipo: Relic 1, Relic 2, Relic 3, Relic 4."
    )

  best = {}

  for r in rows[1:]:
    name = str(r[name_idx] or "").strip()
    if not name:
      continue

    rels = []
    for idx in relic_idxs:
      if idx is None:
        continue
      v = r[idx]
      v = str(v or "").strip()
      if v:
        rels.append(v)

    # guarda solo si tiene al menos 1 reliquia
    if rels:
      best[norm(name)] = rels[:4]

  out_path.parent.mkdir(parents=True, exist_ok=True)
  out_path.write_text(json.dumps(best, ensure_ascii=False, indent=2), encoding="utf-8")
  print(f"OK: generado {out_path} con {len(best)} miscrits.")


if __name__ == "__main__":
  main()
